import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config
from models.database import CraftingStation, GunCode, Weapon, WeaponCategory, get_session

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def export_all_to_excel():
    wb = Workbook()
    session = get_session()
    try:
        ws1 = wb.active
        ws1.title = "武器列表"
        headers = ["武器名称", "分类", "改枪码数量", "来源链接"]
        ws1.append(headers)
        _style_header(ws1, len(headers))
        weapons = session.query(Weapon).order_by(Weapon.name).all()
        for w in weapons:
            cat_name = w.category.name if w.category else "-"
            ws1.append([w.name, cat_name, w.code_count, w.source_url or "-"])
        _auto_width(ws1)

        ws2 = wb.create_sheet("改枪码汇总")
        headers = ["武器", "改枪码", "描述", "价值", "复制次数", "点赞", "游戏模式"]
        ws2.append(headers)
        _style_header(ws2, len(headers))
        codes = (
            session.query(GunCode, Weapon.name)
            .join(Weapon, GunCode.weapon_id == Weapon.id)
            .order_by(Weapon.name, GunCode.copy_count.desc())
            .all()
        )
        for code, weapon_name in codes:
            ws2.append([weapon_name, code.code, code.description or "-", code.value or "-",
                        code.copy_count, code.likes, code.game_mode or "烽火地带"])
        _auto_width(ws2)

        ws3 = wb.create_sheet("特勤处推荐")
        headers = ["工作台", "推荐物品", "小时利润"]
        ws3.append(headers)
        _style_header(ws3, len(headers))
        stations = session.query(CraftingStation).order_by(CraftingStation.id).all()
        for s in stations:
            ws3.append([s.station_name, s.item_name, s.hourly_profit])
        _auto_width(ws3)
    finally:
        session.close()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"三角洲改枪码_{timestamp}.xlsx"
    filepath = os.path.join(Config.EXPORT_DIR, filename)
    os.makedirs(Config.EXPORT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath


def export_weapon_to_excel(weapon_slug):
    session = get_session()
    try:
        weapon = session.query(Weapon).filter_by(slug=weapon_slug).first()
        if not weapon:
            return None
        wb = Workbook()
        ws = wb.active
        ws.title = weapon.name
        headers = ["改枪码", "描述", "价值", "复制次数", "点赞", "游戏模式"]
        ws.append(headers)
        _style_header(ws, len(headers))
        codes = session.query(GunCode).filter_by(weapon_id=weapon.id).order_by(GunCode.copy_count.desc()).all()
        for c in codes:
            ws.append([c.code, c.description or "-", c.value or "-", c.copy_count, c.likes, c.game_mode or "烽火地带"])
        _auto_width(ws)
    finally:
        session.close()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{weapon_slug}_{timestamp}.xlsx"
    filepath = os.path.join(Config.EXPORT_DIR, filename)
    os.makedirs(Config.EXPORT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath
